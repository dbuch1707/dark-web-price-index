"""
Dark Web Price Index — 2024 Interpolation Script
==================================================
Generates estimated 2024 price values using linear interpolation
between 2023 actuals and August 2025 actuals.

Method:
    2024_value = (2023_value + 2025_value) / 2

All interpolated rows are clearly flagged:
    Is_Estimated = "Yes"
    Price_Note   = "Estimated – linear interpolation 2023→2025"

Usage:
    python 02_interpolate_2024.py

Input:
    Dark_Web_Price_Index_All_Years.csv   (your existing dataset)
    OR scraped_raw.csv                   (from 01_scrape_dark_web_prices.py)

Output:
    Dark_Web_Price_Index_With_2024.csv   — full dataset with 2024 added
    Dark_Web_Price_Index_With_2024.xlsx  — Excel version, color-coded
    interpolation_report.txt             — audit trail of every 2024 estimate
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")


# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_FILE  = "Dark_Web_Price_Index_All_Years.csv"   # change if using scraped_raw.csv
OUTPUT_CSV  = "Dark_Web_Price_Index_With_2024.csv"
OUTPUT_XLSX = "Dark_Web_Price_Index_With_2024.xlsx"
REPORT_FILE = "interpolation_report.txt"

# Cells shaded yellow in Excel = interpolated rows
ESTIMATED_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
ACTUAL_FILL    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_FILL    = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
HEADER_FONT    = Font(color="FFFFFF", bold=True, name="Arial", size=10)


# ── LOAD DATA ─────────────────────────────────────────────────────────────────

def load_data(filepath: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(filepath)
        print(f"✓ Loaded {len(df)} rows from {filepath}")
        print(f"  Years present: {sorted(df['Year'].unique())}")
        print(f"  Categories: {sorted(df['Category'].unique())}")
        return df
    except FileNotFoundError:
        print(f"✗ File not found: {filepath}")
        print(f"  Run 01_scrape_dark_web_prices.py first, or check the filename.")
        raise


# ── INTERPOLATION ─────────────────────────────────────────────────────────────

def interpolate_2024(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    """
    For each (Category, Product, Country, Organization) group that has
    both a 2023 row and a 2025 row, generate a 2024 estimate.

    Returns:
        df_2024  — DataFrame of new 2024 rows only
        audit    — list of dicts describing each interpolation decision
    """
    rows_2023 = df[df["Year"] == 2023].copy()
    rows_2025 = df[df["Year"] == 2025].copy()

    # If no 2025 data, try to use 2025 rows from any source
    if rows_2025.empty:
        print("  ⚠ No 2025 rows found — using latest available year as upper bound")
        latest_year = df["Year"].max()
        rows_2025   = df[df["Year"] == latest_year].copy()

    # Join keys — match on Product + Category (most reliable)
    join_keys = ["Category", "Product"]

    merged = pd.merge(
        rows_2023[join_keys + ["Price_Low_USD","Price_High_USD","Price_Mid_USD","Country","Organization","Source"]],
        rows_2025[join_keys + ["Price_Low_USD","Price_High_USD","Price_Mid_USD","Source"]],
        on=join_keys,
        suffixes=("_2023","_2025")
    )

    print(f"\n  Found {len(merged)} matched pairs (2023 ↔ 2025)")

    new_rows = []
    audit    = []

    for _, row in merged.iterrows():
        lo_2023 = row["Price_Low_USD_2023"]
        hi_2023 = row["Price_High_USD_2023"]
        lo_2025 = row["Price_Low_USD_2025"]
        hi_2025 = row["Price_High_USD_2025"]

        # Linear interpolation: midpoint between the two years
        lo_2024  = round((lo_2023 + lo_2025) / 2, 2)
        hi_2024  = round((hi_2023 + hi_2025) / 2, 2)
        mid_2024 = round((lo_2024 + hi_2024) / 2, 2)

        # Direction of price change for audit
        direction = "↑" if lo_2025 > lo_2023 else ("↓" if lo_2025 < lo_2023 else "→")

        new_rows.append({
            "Year":          2024,
            "Category":      row["Category"],
            "Product":       row["Product"],
            "Country":       row.get("Country", "United States"),
            "Organization":  row.get("Organization", "Various"),
            "Price_Low_USD": lo_2024,
            "Price_High_USD":hi_2024,
            "Price_Mid_USD": mid_2024,
            "Price_Note":    "Estimated – linear interpolation 2023→2025",
            "Source":        "Interpolated",
            "Is_Estimated":  "Yes",
        })

        audit.append({
            "Product":       row["Product"],
            "Category":      row["Category"],
            "2023_Low":      lo_2023,
            "2023_High":     hi_2023,
            "2024_Low_est":  lo_2024,
            "2024_High_est": hi_2024,
            "2025_Low":      lo_2025,
            "2025_High":     hi_2025,
            "Direction":     direction,
            "Pct_Change_2023_to_2025": round(((lo_2025 - lo_2023) / lo_2023 * 100), 1) if lo_2023 > 0 else "N/A",
        })

    # Also carry forward 2025-only items (no 2023 match) with a hold-flat estimate
    products_interpolated = set(merged["Product"])
    for _, row in rows_2025.iterrows():
        if row["Product"] not in products_interpolated:
            # Hold at 2025 value — best estimate when no 2023 baseline
            lo_2024  = row["Price_Low_USD"]
            hi_2024  = row["Price_High_USD"]
            mid_2024 = round((lo_2024 + hi_2024) / 2, 2)

            new_rows.append({
                "Year":          2024,
                "Category":      row["Category"],
                "Product":       row["Product"],
                "Country":       row.get("Country","United States"),
                "Organization":  row.get("Organization","Various"),
                "Price_Low_USD": lo_2024,
                "Price_High_USD":hi_2024,
                "Price_Mid_USD": mid_2024,
                "Price_Note":    "Estimated – held from 2025 (no 2023 baseline available)",
                "Source":        "Interpolated",
                "Is_Estimated":  "Yes",
            })

            audit.append({
                "Product":       row["Product"],
                "Category":      row["Category"],
                "2023_Low":      "N/A",
                "2023_High":     "N/A",
                "2024_Low_est":  lo_2024,
                "2024_High_est": hi_2024,
                "2025_Low":      lo_2024,
                "2025_High":     hi_2024,
                "Direction":     "→ (hold)",
                "Pct_Change_2023_to_2025": "N/A",
            })

    df_2024 = pd.DataFrame(new_rows)
    print(f"  ✓ Generated {len(df_2024)} 2024 estimated rows")
    return df_2024, audit


# ── WRITE AUDIT REPORT ────────────────────────────────────────────────────────

def write_report(audit: list[dict], output_path: str):
    lines = [
        "=" * 70,
        "INTERPOLATION AUDIT REPORT",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 70,
        "",
        "METHOD: Linear interpolation — 2024 = (2023 + 2025) / 2",
        "        For items with no 2023 baseline, 2025 value is held flat.",
        "",
        "IMPORTANT: These are estimates, not primary research.",
        "           Label clearly in any published visualization.",
        "",
        "-" * 70,
        f"{'Product':<42} {'Dir':>3}  {'2023 Low':>10}  {'2024 Est':>10}  {'2025 Low':>10}  {'% Change':>9}",
        "-" * 70,
    ]

    for a in sorted(audit, key=lambda x: x["Category"]):
        lo_23 = f"${a['2023_Low']:,.0f}" if a["2023_Low"] != "N/A" else "N/A"
        lo_24 = f"${a['2024_Low_est']:,.0f}"
        lo_25 = f"${a['2025_Low']:,.0f}" if a["2025_Low"] != "N/A" else "N/A"
        pct   = f"{a['Pct_Change_2023_to_2025']}%" if a["Pct_Change_2023_to_2025"] != "N/A" else "N/A"
        prod  = a["Product"][:40]
        lines.append(f"{prod:<42} {a['Direction']:>3}  {lo_23:>10}  {lo_24:>10}  {lo_25:>10}  {pct:>9}")

    lines += [
        "-" * 70,
        f"\nTotal 2024 rows generated: {len(audit)}",
    ]

    with open(output_path, "w") as f:
        f.write("\n".join(lines))

    print(f"✓ Audit report saved: {output_path}")


# ── WRITE EXCEL WITH COLOR CODING ─────────────────────────────────────────────

def write_excel(df: pd.DataFrame, output_path: str):
    """
    Save to Excel with:
    - Black header row
    - Yellow fill on all 2024 (estimated) rows
    - White fill on all actual rows
    - Auto column widths
    """
    df.to_excel(output_path, index=False, sheet_name="All_Prices")

    wb = load_workbook(output_path)
    ws = wb["All_Prices"]

    # Style header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style data rows
    for row in ws.iter_rows(min_row=2):
        year_val      = row[0].value   # Year column
        is_estimated  = str(row[10].value if len(row) > 10 else "").strip()
        fill = ESTIMATED_FILL if (year_val == 2024 or is_estimated == "Yes") else ACTUAL_FILL
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

    # Format price columns as currency
    price_cols = [6, 7, 8]  # Price_Low, Price_High, Price_Mid (1-indexed)
    for row in ws.iter_rows(min_row=2):
        for col_idx in price_cols:
            if col_idx <= len(row):
                row[col_idx - 1].number_format = '$#,##0.00'

    # Auto column widths
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10
        ) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 52)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"✓ Excel saved: {output_path} (yellow = estimated 2024 rows)")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Dark Web Price Index — 2024 Interpolation")
    print(f"Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Load existing data
    df = load_data(INPUT_FILE)

    # Check we have what we need
    years = sorted(df["Year"].unique())
    if 2024 in years:
        print("\n⚠ 2024 data already present in dataset.")
        print("  Remove existing 2024 rows first, or use a different input file.")

    print(f"\n→ Interpolating 2024 values...")
    df_2024, audit = interpolate_2024(df)

    if df_2024.empty:
        print("✗ No 2024 rows generated — check your input data has 2023 and 2025 rows.")
        return

    # Combine with existing data
    df_full = pd.concat([df, df_2024], ignore_index=True)
    df_full = df_full.sort_values(["Year","Category","Price_Mid_USD"], ascending=[True,True,False])
    df_full = df_full.reset_index(drop=True)

    # Save outputs
    print(f"\n→ Saving outputs...")
    df_full.to_csv(OUTPUT_CSV, index=False)
    print(f"✓ CSV saved: {OUTPUT_CSV} — {len(df_full)} total rows")

    write_excel(df_full, OUTPUT_XLSX)
    write_report(audit, REPORT_FILE)

    # Summary
    print("\n── Final dataset summary ──")
    summary = df_full.groupby(["Year","Is_Estimated"])["Product"].count()
    print(summary.to_string())

    print("\n── 2024 interpolated values (sample) ──")
    sample_cols = ["Category","Product","Price_Low_USD","Price_Mid_USD","Price_High_USD"]
    print(df_2024[sample_cols].head(10).to_string(index=False))

    print("\n✓ Done. Import Dark_Web_Price_Index_With_2024.xlsx into Tableau.")
    print("  Yellow rows = 2024 estimates. Add Is_Estimated to your tooltip for transparency.")


if __name__ == "__main__":
    main()
